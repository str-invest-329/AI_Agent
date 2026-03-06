"""
build_xlsx.py — Generate SNDK_financials.xlsx from SNDK_financials.json
Pivot-table style: rows = metrics, columns = periods
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(DIR, "SNDK_financials.json")
XLSX_PATH = os.path.join(DIR, "SNDK_financials.xlsx")

# Styles
H_FILL = PatternFill("solid", fgColor="1F4E79")
SEC_FILL = PatternFill("solid", fgColor="D6E4F0")
TOT_FILL = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL = PatternFill("solid", fgColor="F2F8FD")
W_FILL = PatternFill("solid", fgColor="FFFFFF")
DERIVED_FILL = PatternFill("solid", fgColor="E2EFDA")

H_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
T_FONT = Font(name="Calibri", bold=True, color="1F4E79", size=13)
B_FONT = Font(name="Calibri", size=10)
BD_FONT = Font(name="Calibri", bold=True, size=10)
TOT_FONT = Font(name="Calibri", bold=True, color="7B3F00", size=10)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CA = Alignment(horizontal="center", vertical="center", wrap_text=True)
RA = Alignment(horizontal="right", vertical="center")
LA = Alignment(horizontal="left", vertical="center")

PERIOD_ENDS = {
    "FY2025_START": "2024-06-28",
    "Q1_FY2025": "2024-09-27",
    "Q2_FY2025": "2024-12-27",
    "Q3_FY2025": "2025-03-28",
    "Q4_FY2025": "2025-06-27",
    "Q1_FY2026": "2025-10-03",
    "Q2_FY2026": "2026-01-02",
}


def cs(c, value, font=B_FONT, fill=W_FILL, align=RA, num_fmt=None):
    c.value = value
    c.font = font
    c.fill = fill
    c.alignment = align
    c.border = BORDER
    if num_fmt and isinstance(value, (int, float)):
        c.number_format = num_fmt


def title_row(ws, row, text, ncols):
    ec = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{ec}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = T_FONT
    c.fill = SEC_FILL
    c.alignment = CA
    ws.row_dimensions[row].height = 22


def write_header_row(ws, row, periods, label=""):
    cs(ws.cell(row=row, column=1), label, font=H_FONT, fill=H_FILL, align=CA)
    for i, p in enumerate(periods, 2):
        end = PERIOD_ENDS.get(p, "")
        cs(ws.cell(row=row, column=i), f"{p}\n{end}", font=H_FONT, fill=H_FILL, align=CA)
    ws.row_dimensions[row].height = 36


def write_section_header(ws, row, label, ncols):
    cs(ws.cell(row=row, column=1), f"  {label}", font=BD_FONT, fill=SEC_FILL, align=LA)
    for col in range(2, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.value = ""
        c.fill = SEC_FILL
        c.border = BORDER


def write_metric_row(ws, row, label, vals, periods, is_total=False, num_fmt="#,##0", pct=False, derived_periods=None):
    fill = TOT_FILL if is_total else (ALT_FILL if row % 2 == 0 else W_FILL)
    font = TOT_FONT if is_total else B_FONT
    bold_font = TOT_FONT if is_total else BD_FONT

    cs(ws.cell(row=row, column=1), label, font=bold_font if is_total else B_FONT, fill=fill, align=LA)
    for i, p in enumerate(periods, 2):
        v = vals.get(p)
        display = v if v is not None else "—"
        nf = ("0.0%" if pct else num_fmt) if isinstance(v, (int, float)) else None
        cell_fill = DERIVED_FILL if (derived_periods and p in derived_periods and v is not None) else fill
        cs(ws.cell(row=row, column=i), display, font=font if is_total else B_FONT, fill=cell_fill, num_fmt=nf)


def set_widths(ws, label_w, data_w, ncols):
    ws.column_dimensions["A"].width = label_w
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = data_w


def build():
    with open(JSON_PATH, "r") as f:
        data = json.load(f)

    periods_is = data["metadata"]["periods_income_statement"]
    periods_bs = data["metadata"]["periods_balance_sheet"]
    ncols_is = len(periods_is) + 1
    ncols_bs = len(periods_bs) + 1
    derived = {"Q4_FY2025"}  # highlight derived periods

    wb = Workbook()

    # ═══ Overview ═══
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "Sandisk Corporation (SNDK) — Financial Timeseries"
    c.font = T_FONT
    c.fill = SEC_FILL
    c.alignment = CA
    ws.row_dimensions[1].height = 28

    r = 2
    for k, v in [
        ("Company", data["metadata"]["company"]),
        ("Ticker", f"{data['metadata']['ticker']} / {data['metadata']['exchange']}"),
        ("CIK", data["metadata"]["cik"]),
        ("Fiscal Year End", data["metadata"]["fiscal_year_end"]),
        ("Currency/Unit", f"{data['metadata']['currency']} / {data['metadata']['unit']}"),
        ("Data Source", data["metadata"]["data_source"]),
        ("Last Updated", data["metadata"]["last_updated"]),
        ("IS Periods", " | ".join(periods_is)),
        ("BS Periods", " | ".join(periods_bs)),
        ("Notes", data["metadata"]["notes"]),
    ]:
        for col, val in [(1, k), (2, v)]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = BD_FONT if col == 1 else B_FONT
            c.fill = ALT_FILL if r % 2 == 0 else W_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r += 1

    r += 1
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(row=r, column=1, value="Filing Registry")
    c.font = H_FONT
    c.fill = H_FILL
    c.alignment = CA
    r += 1
    for p, info in data["filings"].items():
        for col, val in [(1, p), (2, f"{info['form']} | Period end: {info['period_end']} | Filed: {info['filing_date']} | Accession: {info['accession_number']}")]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = BD_FONT if col == 1 else B_FONT
            c.fill = ALT_FILL if r % 2 == 0 else W_FILL
            c.border = BORDER
            c.alignment = LA
        r += 1

    # ═══ Income Statement ═══
    ws2 = wb.create_sheet("Income Statement")
    ws2.sheet_view.showGridLines = False
    set_widths(ws2, 38, 16, ncols_is)
    r = 1
    title_row(ws2, r, "Consolidated Statements of Operations (USD Millions, except per share)", ncols_is)
    r += 1
    write_header_row(ws2, r, periods_is)
    r += 1

    IS = data["income_statement"]
    is_rows = [
        ("Revenue", "revenue", True, "#,##0", False),
        ("Cost of Goods Sold", "cost_of_goods_sold", False, "#,##0", False),
        ("Gross Profit", "gross_profit", True, "#,##0", False),
        ("  Gross Margin %", "gross_margin_pct", False, "0.0%", True),
        ("R&D Expense", "research_and_development", False, "#,##0", False),
        ("SG&A Expense", "selling_general_administrative", False, "#,##0", False),
        ("Restructuring Charges", "restructuring_charges", False, "#,##0", False),
        ("Operating Income (Loss)", "operating_income", True, "#,##0", False),
        ("  Operating Margin %", "operating_margin_pct", False, "0.0%", True),
        ("Interest Expense", "interest_expense", False, "#,##0", False),
        ("Interest Income", "interest_income", False, "#,##0", False),
        ("Other Non-op Income (Expense)", "other_nonoperating_income_expense", False, "#,##0", False),
        ("Non-op Income (Expense) Total", "nonoperating_income_expense_total", False, "#,##0", False),
        ("Income Before Taxes", "income_before_taxes", True, "#,##0", False),
        ("Income Tax Expense", "income_tax_expense", False, "#,##0", False),
        ("  Effective Tax Rate", "effective_tax_rate", False, "0.0%", True),
        ("Equity Method Investments", "equity_method_investments", False, "#,##0", False),
        ("Net Income (Loss)", "net_income", True, "#,##0", False),
        ("  Net Margin %", "net_margin_pct", False, "0.0%", True),
        ("EPS — Basic", "eps_basic", False, "$#,##0.00", False),
        ("EPS — Diluted", "eps_diluted", True, "$#,##0.00", False),
        ("Shares Basic (M)", "shares_basic_millions", False, "#,##0.0", False),
        ("Shares Diluted (M)", "shares_diluted_millions", False, "#,##0.0", False),
    ]
    for label, key, is_tot, nf, pct in is_rows:
        write_metric_row(ws2, r, label, IS.get(key, {}), periods_is,
                         is_total=is_tot, num_fmt=nf, pct=pct, derived_periods=derived)
        r += 1

    # ═══ Balance Sheet ═══
    ws3 = wb.create_sheet("Balance Sheet")
    ws3.sheet_view.showGridLines = False
    set_widths(ws3, 40, 16, ncols_bs)
    r = 1
    title_row(ws3, r, "Consolidated Balance Sheets (USD Millions, as of period end)", ncols_bs)
    r += 1
    write_header_row(ws3, r, periods_bs)
    r += 1

    BS = data["balance_sheet"]
    for sec_name, sec_key, items in [
        ("ASSETS", "assets", [
            ("Cash & Cash Equivalents", "cash_and_cash_equivalents", False),
            ("Accounts Receivable", "accounts_receivable", False),
            ("Inventories", "inventories", False),
            ("Other Current Assets", "other_current_assets", False),
            ("Total Current Assets", "total_current_assets", True),
            ("PP&E (net)", "property_plant_equipment_net", False),
            ("Operating Lease ROU Asset", "operating_lease_rou_asset", False),
            ("Goodwill", "goodwill", False),
            ("Deferred Tax Assets (net)", "deferred_tax_assets_net", False),
            ("Other Noncurrent Assets", "other_noncurrent_assets", False),
            ("Total Assets", "total_assets", True),
        ]),
        ("LIABILITIES", "liabilities", [
            ("Accrued Liabilities", "accrued_liabilities_current", False),
            ("Employee Liabilities", "employee_related_liabilities_current", False),
            ("Current Debt", "current_debt", False),
            ("Operating Lease (Current)", "operating_lease_liability_current", False),
            ("Total Current Liabilities", "total_current_liabilities", True),
            ("Long-term Debt", "long_term_debt_noncurrent", False),
            ("Operating Lease (Noncurrent)", "operating_lease_liability_noncurrent", False),
            ("Other Noncurrent Liabilities", "other_noncurrent_liabilities", False),
            ("Total Liabilities", "total_liabilities", True),
        ]),
        ("EQUITY", "equity", [
            ("Common Stock", "common_stock", False),
            ("Additional Paid-in Capital", "additional_paid_in_capital", False),
            ("Retained Earnings (Deficit)", "retained_earnings", False),
            ("AOCI", "aoci", False),
            ("Total Equity", "total_equity", True),
            ("Total Liabilities & Equity", "total_liabilities_and_equity", True),
        ]),
    ]:
        write_section_header(ws3, r, sec_name, ncols_bs)
        r += 1
        for label, key, is_tot in items:
            write_metric_row(ws3, r, label, BS[sec_key].get(key, {}), periods_bs, is_total=is_tot)
            r += 1

    # ═══ Cash Flow ═══
    ws4 = wb.create_sheet("Cash Flow")
    ws4.sheet_view.showGridLines = False
    set_widths(ws4, 44, 16, ncols_is)
    r = 1
    title_row(ws4, r, "Consolidated Statements of Cash Flows (USD Millions)", ncols_is)
    r += 1
    write_header_row(ws4, r, periods_is)
    r += 1

    CF = data["cash_flow_statement"]
    for sec_name, sec_key, items in [
        ("OPERATING ACTIVITIES", "operating_activities", [
            ("Net Income", "net_income", False),
            ("Depreciation & Amortization", "depreciation_and_amortization", False),
            ("Share-based Compensation", "share_based_compensation", False),
            ("Deferred Income Tax", "deferred_income_tax", False),
            ("Goodwill Impairment", "goodwill_impairment", False),
            ("Other Asset Impairment", "other_asset_impairment", False),
            ("Gain/Loss on Sale of Business", "gain_loss_on_sale_of_business", False),
            ("Change in Receivables", "change_in_receivables", False),
            ("Change in Inventories", "change_in_inventories", False),
            ("Change in Accounts Payable", "change_in_accounts_payable", False),
            ("Change in Accrued Liabilities", "change_in_accrued_liabilities", False),
            ("Change in Employee Liabilities", "change_in_employee_liabilities", False),
            ("Other Operating", "other_operating", False),
            ("Net Cash from Operating", "net_cash_from_operating", True),
        ]),
        ("INVESTING ACTIVITIES", "investing_activities", [
            ("Capital Expenditures", "capital_expenditures", False),
            ("Proceeds from Sale of Business", "proceeds_from_sale_of_business", False),
            ("Proceeds from Sale of PP&E", "proceeds_from_sale_of_ppe", False),
            ("Net Cash from Investing", "net_cash_from_investing", True),
        ]),
        ("FINANCING ACTIVITIES", "financing_activities", [
            ("Proceeds from Debt", "proceeds_from_debt", False),
            ("Repayments of Debt", "repayments_of_debt", False),
            ("Proceeds from Equity", "proceeds_from_equity", False),
            ("Tax Withholding Payments", "tax_withholding_payments", False),
            ("Net Cash from Financing", "net_cash_from_financing", True),
        ]),
    ]:
        write_section_header(ws4, r, sec_name, ncols_is)
        r += 1
        for label, key, is_tot in items:
            write_metric_row(ws4, r, label, CF[sec_key].get(key, {}), periods_is,
                             is_total=is_tot, derived_periods=derived)
            r += 1

    write_section_header(ws4, r, "SUMMARY", ncols_is)
    r += 1
    for label, key, is_tot in [
        ("FX Effect on Cash", "fx_effect_on_cash", False),
        ("Net Change in Cash", "net_change_in_cash", False),
        ("Ending Cash", "ending_cash", False),
        ("Free Cash Flow (OCF - CapEx)", "free_cash_flow", True),
    ]:
        write_metric_row(ws4, r, label, CF.get(key, {}), periods_is,
                         is_total=is_tot, derived_periods=derived)
        r += 1

    # ═══ Financial Ratios ═══
    ws5 = wb.create_sheet("Financial Ratios")
    ws5.sheet_view.showGridLines = False
    set_widths(ws5, 32, 16, ncols_bs)
    r = 1
    title_row(ws5, r, "Financial Ratios (computed from balance sheet)", ncols_bs)
    r += 1
    write_header_row(ws5, r, periods_bs)
    r += 1

    ratios = data.get("financial_ratios", {})
    ratio_metrics = [
        ("Current Ratio", "current_ratio", True, "#,##0.00"),
        ("Quick Ratio", "quick_ratio", False, "#,##0.00"),
        ("Debt-to-Equity", "debt_to_equity", False, "#,##0.00"),
        ("Net Debt-to-Equity", "net_debt_to_equity", False, "#,##0.00"),
        ("Equity Ratio", "equity_ratio", False, "0.0%"),
    ]
    for label, key, is_tot, nf in ratio_metrics:
        vals = {p: r_dict.get(key) for p, r_dict in ratios.items()}
        pct = "%" in nf
        write_metric_row(ws5, r, label, vals, periods_bs, is_total=is_tot, num_fmt=nf, pct=pct)
        r += 1

    # ═══ Long Format ═══
    ws6 = wb.create_sheet("Long Format (DB)")
    ws6.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [14, 14, 28, 42, 16, 14]):
        ws6.column_dimensions[col].width = w
    title_row(ws6, 1, "Long Format — Tidy Data for Database / Analytics Import", ncols=6)
    for ci, h in enumerate(["period", "period_end", "statement", "metric", "value", "unit"], 1):
        c = ws6.cell(row=2, column=ci, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = CA
        c.border = BORDER

    for i, row_data in enumerate(data.get("long_format", []), 3):
        fill = ALT_FILL if i % 2 == 0 else W_FILL
        for ci, key in enumerate(["period", "period_end", "statement", "metric", "value", "unit"], 1):
            c = ws6.cell(row=i, column=ci, value=row_data.get(key))
            c.font = B_FONT
            c.fill = fill
            c.border = BORDER
            c.alignment = RA if ci == 5 else LA
            if ci == 5 and isinstance(row_data.get("value"), (int, float)):
                c.number_format = "#,##0.####"

    # ═══ Notable Events ═══
    ws7 = wb.create_sheet("Notable Events")
    ws7.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [14, 16, 14, 80]):
        ws7.column_dimensions[col].width = w
    title_row(ws7, 1, "Notable Events", ncols=4)
    for ci, h in enumerate(["period", "category", "date", "description"], 1):
        c = ws7.cell(row=2, column=ci, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = CA
        c.border = BORDER
    for i, ev in enumerate(data.get("notable_events", []), 3):
        fill = ALT_FILL if i % 2 == 0 else W_FILL
        for ci, key in enumerate(["period", "category", "date", "description"], 1):
            c = ws7.cell(row=i, column=ci, value=ev.get(key))
            c.font = B_FONT
            c.fill = fill
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(ci == 4))
        ws7.row_dimensions[i].height = 36

    wb.save(XLSX_PATH)
    print(f"✓ XLSX saved: {XLSX_PATH}")
    print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
    print(f"  Long Format rows: {len(data.get('long_format', []))}")


if __name__ == "__main__":
    build()
